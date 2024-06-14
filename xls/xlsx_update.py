import openpyxl

file = openpyxl.load_workbook('produceSales.xlsx')
sheet = file['Sheet']
sheet.freeze_panes = 'A2'

Price_update = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for row in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row, column=1).value
    if produce_name in Price_update:
        print(f'Updating {produce_name}')
        sheet.cell(row=row, column=2).value = Price_update[produce_name]

file.save('updatedsales.xlsx')
